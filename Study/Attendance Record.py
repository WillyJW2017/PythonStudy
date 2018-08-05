# 整理考勤记录：找出每天员工的上班打卡时间和下班打卡时间
# 计算出上班时间 =下班打卡时间 - 上班打卡时间
# 最后再计算出实际上班时间 = 上班时间 - 1.5H
import openpyxl
import datetime


filePath = 'D:\\Willy\\Report\\2.xlsx'
sheetName = 'Sheet0'
workbook = openpyxl.load_workbook(filePath)
sheet = workbook[sheetName]

start_row = 2
start_num = 2
full_data_dict = {}
date_dict = {}
# date_dict里面存放每天所有员工总打卡次数：{'07/01/2018': 38, '07/02/2018': 68, '07/03/2018': 90, '07/04/2018': 95}
for i in range(2, sheet.max_row+1):
    date_value = sheet.cell(row=i, column=3).value
    date_dict.setdefault(date_value, 0)
    date_dict[date_value] += 1

date_list = date_dict.keys()
for date_v in date_list:
    # full_name_dict存放了一天之内每个员工的总打卡次数
    full_name_dict = {}
    end_row = date_dict[date_v] + start_row
    for m in range(start_row, end_row):
        fname = sheet.cell(row=m, column=1).value
        lname = sheet.cell(row=m, column=2).value

        full_name = fname + ' ' + lname
        full_name_dict.setdefault(full_name, 0)
        full_name_dict[full_name] += 1

    # full_data_dict：key为日期，value为 员工姓名:打卡次数
    # {'07/01/2018': {'BEN HO': 4, 'CACY XU': 4}, '07/02/2018': {'AMANDA MEI': 7, 'AMY RUI': 6}}
    full_data_dict[date_v] = full_name_dict
    start_row = end_row

    full_name_list = full_name_dict.keys()
    for full_name in full_name_list:
        end_num = full_name_dict[full_name] + start_num
        min_time_num = start_num    # 临时最小值行数
        max_time_num = start_num    # 临时最大值行数

        for x in range(start_num, end_num, 2):
            fir_time = sheet.cell(row=x, column=4).value
            if x == end_num - 1:
                sec_time = sheet.cell(row=x, column=4).value
            else:
                sec_time = sheet.cell(row=x+1, column=4).value

            fir_time = datetime.datetime.strptime(fir_time, '%H:%M:%S')
            sec_time = datetime.datetime.strptime(sec_time, '%H:%M:%S')

            min_value = sheet.cell(row=min_time_num, column=4).value
            max_value = sheet.cell(row=max_time_num, column=4).value
            min_value = datetime.datetime.strptime(min_value, '%H:%M:%S')
            max_value = datetime.datetime.strptime(max_value, '%H:%M:%S')

            end_num_value = sheet.cell(row=end_num-1, column=4).value
            end_num_value = datetime.datetime.strptime(end_num_value, '%H:%M:%S')
            # 找出最大值和最小值所在的行数
            if fir_time < sec_time:
                if fir_time < min_value:
                    min_time_num = x
                if sec_time > max_value:
                    max_time_num = x+1
            else:
                if sec_time < min_value:
                    min_time_num = x+1
                if fir_time > max_value:
                    max_time_num = x
            # 如果打卡次数为基数就要对最后一个值单独拿出来做一次判断
            if (end_num - start_num) % 2 != 0 :
                if end_num_value < min_value:
                    min_time_num = end_num-1
                if end_num_value > max_value:
                    max_time_num = end_num-1

        min_time_value = sheet.cell(row=min_time_num, column=4).value
        max_time_value = sheet.cell(row=max_time_num, column=4).value

        start_num = end_num

        # 将最后打卡时间写入第一次打卡所在的行数对应的单元格
        sheet.cell(row=min_time_num, column=5, value=max_time_value)

        # 计算每天上班时间：最后打卡时间 - 最早打卡时间
        min_time_value = datetime.datetime.strptime(min_time_value, '%H:%M:%S')
        max_time_value = datetime.datetime.strptime(max_time_value, '%H:%M:%S')
        hours = max_time_value - min_time_value
        sheet.cell(row=min_time_num, column=6, value=hours)

        # 计算实际上班时间: 打卡时间 - 1.5H
        actual_hours = hours + datetime.timedelta(hours=-1.5)
        sheet.cell(row=min_time_num, column=7, value=actual_hours)

workbook.save(filename = filePath)
